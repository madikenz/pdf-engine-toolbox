"""Test office-to-PDF conversion via LibreOffice."""

import json
import shutil
from unittest.mock import patch, AsyncMock

import fitz
import pytest


def _has_libreoffice() -> bool:
    """Check if LibreOffice is available on the system."""
    return shutil.which("libreoffice") is not None or shutil.which("soffice") is not None


@pytest.mark.skipif(not _has_libreoffice(), reason="LibreOffice not installed")
def test_convert_docx_to_pdf(client, auth_headers):
    """A simple DOCX (created via python-docx-like bytes) should convert to PDF."""
    # Create a minimal .docx-like file. Since we can't easily create a real DOCX
    # in a test, we'll use a plain text file with a .txt extension, which
    # LibreOffice can also convert.
    text_content = b"Hello from PDF Engine\nThis is a test document for review.\n"

    body = json.dumps({
        "source_url": "https://example.com/test.txt",
        "filename": "test-report.txt",
    })
    headers = auth_headers("POST", "/convert/from-office", body)

    with patch(
        "app.services.download_service.download_pdf",
        new_callable=AsyncMock,
        return_value=text_content,
    ):
        response = client.post("/convert/from-office", content=body, headers=headers)

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"

    # Verify we got a valid PDF
    doc = fitz.open(stream=response.content, filetype="pdf")
    assert len(doc) >= 1
    text = doc[0].get_text()
    assert "Hello from PDF Engine" in text


@pytest.mark.skipif(not _has_libreoffice(), reason="LibreOffice not installed")
def test_convert_csv_to_pdf(client, auth_headers):
    """A CSV file should convert to a PDF (spreadsheet-style)."""
    csv_content = b"Name,Amount,Date\nW-2 Income,85000,2024-01-31\n1099-INT,1250,2024-01-15\n"

    body = json.dumps({
        "source_url": "https://example.com/data.csv",
        "filename": "client-data.csv",
    })
    headers = auth_headers("POST", "/convert/from-office", body)

    with patch(
        "app.services.download_service.download_pdf",
        new_callable=AsyncMock,
        return_value=csv_content,
    ):
        response = client.post("/convert/from-office", content=body, headers=headers)

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"

    # Verify valid PDF
    doc = fitz.open(stream=response.content, filetype="pdf")
    assert len(doc) >= 1


def test_convert_unsupported_format(client, auth_headers):
    """An unsupported file extension should return an error."""
    body = json.dumps({
        "source_url": "https://example.com/malware.exe",
        "filename": "malware.exe",
    })
    headers = auth_headers("POST", "/convert/from-office", body)

    with patch(
        "app.services.download_service.download_pdf",
        new_callable=AsyncMock,
        return_value=b"MZ\x00\x00",
    ):
        response = client.post("/convert/from-office", content=body, headers=headers)

    # Should fail with PDF_CORRUPT error (unsupported format)
    assert response.status_code == 422
    data = response.json()
    assert data["success"] is False
    assert "Unsupported" in data["error"]["message"]


def test_large_office_file_becomes_background_task(client, auth_headers):
    """Files > 5 MB should be processed as a background task."""
    # Create a file larger than 5 MB
    large_content = b"x" * (6 * 1024 * 1024)  # 6 MB of data

    body = json.dumps({
        "source_url": "https://example.com/big-report.xlsx",
        "filename": "big-report.xlsx",
    })
    headers = auth_headers("POST", "/convert/from-office", body)

    with patch(
        "app.services.download_service.download_pdf",
        new_callable=AsyncMock,
        return_value=large_content,
    ):
        response = client.post("/convert/from-office", content=body, headers=headers)

    assert response.status_code == 200
    data = response.json()
    assert data["success"] is True
    assert "task_id" in data
    assert data["status"] == "pending"
    assert "big-report.xlsx" in data["message"]


def test_small_office_file_is_synchronous(client, auth_headers):
    """Files < 5 MB should be processed synchronously (not background task)."""
    # Small text content (well under 5 MB)
    small_content = b"Name,Amount\nTest,100\n"

    body = json.dumps({
        "source_url": "https://example.com/small.txt",
        "filename": "small.txt",
    })
    headers = auth_headers("POST", "/convert/from-office", body)

    with patch(
        "app.services.download_service.download_pdf",
        new_callable=AsyncMock,
        return_value=small_content,
    ), patch(
        "app.services.pdf_service.office_to_pdf",
        return_value=b"%PDF-1.4 mock content",
    ):
        response = client.post("/convert/from-office", content=body, headers=headers)

    assert response.status_code == 200
    # Synchronous: returns PDF directly, not a task_id
    assert response.headers["content-type"] == "application/pdf"


def test_spreadsheet_preprocessing():
    """The _prepare_spreadsheet_for_pdf helper should set fit-to-width on xlsx."""
    import os
    import tempfile
    import openpyxl
    from app.services.pdf_service import _prepare_spreadsheet_for_pdf

    with tempfile.TemporaryDirectory() as tmpdir:
        # Create a simple xlsx
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Amount", "Category", "Date", "Notes"])
        ws.append(["Test", "1000", "Income", "2024-01-15", "Sample data"])
        xlsx_path = os.path.join(tmpdir, "test.xlsx")
        wb.save(xlsx_path)

        # Run the preprocessing
        result_path = _prepare_spreadsheet_for_pdf(xlsx_path, ".xlsx", tmpdir)

        # Should return same path (modified in-place)
        assert result_path == xlsx_path

        # Verify the page setup was applied
        wb2 = openpyxl.load_workbook(result_path)
        ws2 = wb2.active
        assert ws2.page_setup.fitToWidth == 1
        assert ws2.page_setup.fitToHeight == 0
        assert ws2.page_setup.orientation == "landscape"


def test_csv_preprocessing_creates_xlsx():
    """CSV files should be converted to xlsx with fit-to-width settings."""
    import os
    import tempfile
    import openpyxl
    from app.services.pdf_service import _prepare_spreadsheet_for_pdf

    with tempfile.TemporaryDirectory() as tmpdir:
        # Write a CSV file
        csv_path = os.path.join(tmpdir, "data.csv")
        with open(csv_path, "w") as f:
            f.write("Name,Amount,Date\nW-2 Income,85000,2024-01-31\n")

        # Run the preprocessing
        result_path = _prepare_spreadsheet_for_pdf(csv_path, ".csv", tmpdir)

        # Should create a new xlsx file
        assert result_path.endswith(".xlsx")
        assert os.path.exists(result_path)

        # Verify the xlsx has fit-to-width settings
        wb = openpyxl.load_workbook(result_path)
        ws = wb.active
        assert ws.page_setup.fitToWidth == 1
        assert ws.page_setup.fitToHeight == 0
        assert ws.page_setup.orientation == "landscape"

        # Verify the data was preserved
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("Name", "Amount", "Date")
        assert rows[1] == ("W-2 Income", "85000", "2024-01-31")
